"""Excel parser for spreadsheet data extraction."""

from pathlib import Path
from typing import Dict, List, Optional, Union

try:
    import pandas as pd
    HAS_PANDAS = True
except ImportError:
    HAS_PANDAS = False

try:
    from openpyxl import load_workbook
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


class ExcelParser:
    """Extract data from Excel files."""

    def __init__(self):
        """Initialize parser."""
        if not HAS_OPENPYXL and not HAS_PANDAS:
            print("Warning: openpyxl/pandas not installed. Install with: pip install openpyxl pandas")

    def parse(self, path: Path) -> Dict[str, Union[List, "pd.DataFrame"]]:
        """
        Parse all sheets from Excel file.

        Args:
            path: Path to Excel file

        Returns:
            Dict mapping sheet name to DataFrame (or list of lists)
        """
        if HAS_PANDAS:
            return self._parse_with_pandas(path)
        elif HAS_OPENPYXL:
            return self._parse_with_openpyxl(path)
        else:
            raise ImportError("openpyxl or pandas required: pip install openpyxl pandas")

    def _parse_with_pandas(self, path: Path) -> Dict[str, "pd.DataFrame"]:
        """Parse using pandas."""
        excel_file = pd.ExcelFile(path)
        sheets = {}

        for sheet_name in excel_file.sheet_names:
            df = pd.read_excel(excel_file, sheet_name=sheet_name)
            sheets[sheet_name] = df

        return sheets

    def _parse_with_openpyxl(self, path: Path) -> Dict[str, List]:
        """Parse using openpyxl (fallback)."""
        wb = load_workbook(path, data_only=True)
        sheets = {}

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            data = []
            for row in ws.iter_rows(values_only=True):
                data.append(list(row))
            sheets[sheet_name] = data

        return sheets

    def get_sheet_names(self, path: Path) -> List[str]:
        """Get list of sheet names."""
        if HAS_PANDAS:
            excel_file = pd.ExcelFile(path)
            return excel_file.sheet_names
        elif HAS_OPENPYXL:
            wb = load_workbook(path, read_only=True)
            return wb.sheetnames
        return []

    def parse_sheet(self, path: Path, sheet_name: str) -> Union[List, "pd.DataFrame"]:
        """
        Parse a specific sheet.

        Args:
            path: Path to Excel file
            sheet_name: Name of sheet to parse

        Returns:
            DataFrame or list of lists
        """
        sheets = self.parse(path)
        return sheets.get(sheet_name)

    def find_table_region(
        self,
        data: Union[List, "pd.DataFrame"],
        header_keywords: List[str] = None,
    ) -> Optional[Union[List, "pd.DataFrame"]]:
        """
        Find a table region within a sheet based on header keywords.

        Args:
            data: Sheet data (DataFrame or list)
            header_keywords: Keywords to identify header row

        Returns:
            Extracted table region
        """
        if header_keywords is None:
            header_keywords = ["합계", "total", "금액", "amount", "견적"]

        if HAS_PANDAS and isinstance(data, pd.DataFrame):
            # Find row containing keywords
            for idx, row in data.iterrows():
                row_str = " ".join(str(v).lower() for v in row.values if v is not None)
                if any(kw.lower() in row_str for kw in header_keywords):
                    return data.iloc[idx:].reset_index(drop=True)
            return data

        # List-based approach
        for i, row in enumerate(data):
            if row:
                row_str = " ".join(str(v).lower() for v in row if v is not None)
                if any(kw.lower() in row_str for kw in header_keywords):
                    return data[i:]

        return data


# Singleton instance
_parser = None

def get_excel_parser() -> ExcelParser:
    """Get singleton ExcelParser instance."""
    global _parser
    if _parser is None:
        _parser = ExcelParser()
    return _parser
