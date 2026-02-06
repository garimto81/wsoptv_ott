"""
Slack Lists synchronization script.

Syncs Slack Kanban board (Lists) with local VENDOR-DASHBOARD.md and SLACK-LOG.md.
Also updates pinned messages in the project channel.
"""

import json
import sys
from datetime import datetime
from pathlib import Path
from typing import Optional

# Windows cp949 인코딩 에러 방지
if sys.stdout and hasattr(sys.stdout, "reconfigure"):
    try:
        sys.stdout.reconfigure(encoding="utf-8")
    except Exception:
        pass

# Add C:\claude to path for lib.slack import
_script_dir = Path(__file__).resolve().parent
_wsoptv_root = _script_dir.parents[1]
_claude_root = _script_dir.parents[2]
sys.path.insert(0, str(_claude_root))

from lib.slack import SlackClient, SlackUserClient  # noqa: E402

# Import models from same directory
try:
    from .models import SyncResult
except ImportError:
    from models import SyncResult

# Import attachment/parser/extractor modules
try:
    from .collectors.attachment_downloader import AttachmentDownloader
    from .parsers.pdf_parser import get_pdf_parser
    from .parsers.excel_parser import get_excel_parser
    from .extractors.quote_extractor import get_quote_extractor
except ImportError:
    try:
        from collectors.attachment_downloader import AttachmentDownloader
        from parsers.pdf_parser import get_pdf_parser
        from parsers.excel_parser import get_excel_parser
        from extractors.quote_extractor import get_quote_extractor
    except ImportError:
        AttachmentDownloader = None
        get_pdf_parser = None
        get_excel_parser = None
        get_quote_extractor = None

# QuoteFormatter import
try:
    from .formatters.quote_formatter import QuoteFormatter
except ImportError:
    try:
        from formatters.quote_formatter import QuoteFormatter
    except ImportError:
        QuoteFormatter = None


# Slack Lists Configuration (from slack-kanban-config.md)
LISTS_CONFIG = {
    "list_id": "F0ACQJS6KND",
    "team_id": "T03QGJ73GBB",
    "channel_id": "C09TX3M1J2W",
    "pinned_message_ts": "1770011097.422069",  # 고정 요약 메시지 ts
    "columns": {
        "vendor_name": "Col0ACBLXEU66",  # 업체명
        "status": "Col0ABWF6TEQP",        # 상태
        "quote": "Col0ACQRUJ073",         # 견적
        "last_contact": "Col0ACQRVGHKK",  # 최종 연락
        "next_action": "Col0ABWF0K4DV",   # 다음 액션
        "notes": "Col0ABWF1LL9M",         # 비고
    },
    "status_options": {
        "검토 중": "OptN4JBLQ9C",
        "견적 대기": "Opt2QCAWCYN",
        "협상 중": "OptZSZEAEHJ",
        "보류": "OptBAAGMVM4",
        "RFP 검토": "Opt987UPKAJ",  # 맑음소프트 상태
    },
    "items": {
        "메가존클라우드": "Rec0AC9KQD7TQ",
        "맑음소프트": "Rec0ACBM3P1PC",
        "Brightcove": "Rec0AC9KQQEDC",
        "Vimeo OTT": "Rec0ACQJY2W65",
    },
}


# 상태 우선순위 (숫자가 높을수록 진행된 상태)
STATUS_PRIORITY = {
    "보류": 0,        # 특수 상태 - 다른 상태로의 전환 허용
    "견적 대기": 1,
    "검토 중": 2,
    "RFP 검토": 2,
    "협상 중": 3,
    "계약 진행": 4,
}


class QuoteAnalyzer:
    """첨부파일 AI 분석 기반 견적 추론 파이프라인."""

    def __init__(self, cache_dir: Path = None, gmail_client=None):
        """
        Args:
            cache_dir: 첨부파일 저장 디렉토리
            gmail_client: GmailClient 인스턴스 (없으면 자동 생성)
        """
        self.cache_dir = cache_dir or Path("C:/claude/wsoptv_ott/attachments")
        self.cache_dir.mkdir(parents=True, exist_ok=True)
        self.gmail_client = gmail_client
        self._downloader = None
        self._pdf_parser = None
        self._excel_parser = None
        self._quote_extractor = None

    def _get_downloader(self):
        """AttachmentDownloader lazy init."""
        if self._downloader is None:
            if AttachmentDownloader is None:
                raise ImportError("AttachmentDownloader not available")
            self._downloader = AttachmentDownloader(
                gmail_client=self.gmail_client,
                cache_dir=self.cache_dir,
            )
        return self._downloader

    def _get_pdf_parser(self):
        """PDFParser lazy init."""
        if self._pdf_parser is None:
            if get_pdf_parser is None:
                raise ImportError("PDFParser not available")
            self._pdf_parser = get_pdf_parser()
        return self._pdf_parser

    def _get_excel_parser(self):
        """ExcelParser lazy init."""
        if self._excel_parser is None:
            if get_excel_parser is None:
                raise ImportError("ExcelParser not available")
            self._excel_parser = get_excel_parser()
        return self._excel_parser

    def _get_quote_extractor(self):
        """QuoteExtractor lazy init."""
        if self._quote_extractor is None:
            if get_quote_extractor is None:
                raise ImportError("QuoteExtractor not available")
            self._quote_extractor = get_quote_extractor()
        return self._quote_extractor

    def analyze_quote_emails(
        self,
        emails: list[dict],
        changes: list[dict],
    ) -> dict[str, dict]:
        """
        견적 감지된 이메일의 첨부파일을 분석.

        Args:
            emails: Gmail 원본 이메일 객체 리스트 (gmail_formatted 형식)
            changes: VendorChangeDetector.analyze_gmail_messages() 결과

        Returns:
            vendor별 분석 결과 dict
            {
                "Vimeo OTT": {
                    "files_analyzed": 2,
                    "quote_summary": "$42.5K ~ $388K",
                    "options": [...QuoteOption...],
                    "raw_texts": [...],
                },
                ...
            }
        """
        # 견적 감지된 변경만 필터
        quote_changes = [c for c in changes if c.get("quote_detected")]
        if not quote_changes:
            return {}

        # email_id → email 매핑
        email_map = {}
        for email in emails:
            eid = email.get("id", "")
            if eid:
                email_map[eid] = email

        # vendor별 분석 대상 email 수집
        vendor_emails = {}
        for change in quote_changes:
            vendor = change.get("vendor", "")
            if not vendor:
                continue

            # 같은 vendor의 email 중 quote_detected인 것 찾기
            # change에 직접 email_id가 없으면 subject로 매칭
            matched_email_ids = []
            for eid, email in email_map.items():
                email_vendor = self._detect_vendor_from_email(email)
                if email_vendor == vendor and email.get("has_attachments"):
                    matched_email_ids.append(eid)

            if vendor not in vendor_emails:
                vendor_emails[vendor] = set()
            vendor_emails[vendor].update(matched_email_ids)

        # vendor별 첨부파일 다운로드 및 분석
        results = {}
        for vendor, email_ids in vendor_emails.items():
            vendor_result = {
                "files_analyzed": 0,
                "quote_summary": "N/A",
                "options": [],
                "raw_texts": [],
                "errors": [],
            }

            for email_id in email_ids:
                try:
                    attachments = self._download_attachments(email_id)
                    for att_path in attachments:
                        try:
                            # Excel 파일은 직접 추출, PDF는 파싱 후 추출
                            if att_path.suffix.lower() in (".xlsx", ".xls"):
                                quotes = self._extract_quotes_from_excel(att_path)
                                vendor_result["files_analyzed"] += 1
                                vendor_result["options"].extend(quotes)
                            else:
                                parsed_text = self._parse_file(att_path)
                                if parsed_text:
                                    vendor_result["raw_texts"].append(parsed_text)
                                    vendor_result["files_analyzed"] += 1

                                    # 금액 추출
                                    quotes = self._extract_quotes(parsed_text, vendor)
                                    vendor_result["options"].extend(quotes)
                        except Exception as e:
                            vendor_result["errors"].append(
                                f"Parse error ({att_path.name}): {e}"
                            )
                except Exception as e:
                    vendor_result["errors"].append(
                        f"Download error (email {email_id[:8]}): {e}"
                    )

            # 견적 요약 생성
            vendor_result["quote_summary"] = self._generate_ai_summary(
                vendor=vendor,
                parsed_content="\n---\n".join(vendor_result["raw_texts"]),
                extracted_quotes=vendor_result["options"],
            )

            results[vendor] = vendor_result

        return results

    def _detect_vendor_from_email(self, email: dict) -> Optional[str]:
        """이메일에서 vendor 감지 (VendorChangeDetector 로직 재사용)."""
        sender = email.get("sender", "")
        subject = email.get("subject", "")
        email_domain = sender.split("@")[-1].lower() if "@" in sender else ""

        vendor_domains = {
            "brightcove": "Brightcove",
            "megazone": "메가존클라우드",
            "megazonecloud": "메가존클라우드",
            "mz.co.kr": "메가존클라우드",
            "vimeo": "Vimeo OTT",
            "jrmax": "Vimeo OTT",
            "wecandeo": "맑음소프트 (WECANDEO)",
            "malgum": "맑음소프트 (WECANDEO)",
        }

        for domain_key, vendor_name in vendor_domains.items():
            if domain_key in email_domain or domain_key in subject.lower():
                return vendor_name
        return None

    def _download_attachments(self, email_id: str) -> list[Path]:
        """
        이메일 첨부파일 다운로드 (캐시 활용).

        Args:
            email_id: Gmail message ID

        Returns:
            다운로드된 파일 경로 리스트 (PDF/Excel만)
        """
        downloader = self._get_downloader()
        attachments = downloader.get_quote_attachments(email_id)

        paths = []
        for att in attachments:
            if att.local_path:
                path = Path(att.local_path)
                if path.exists():
                    paths.append(path)

        return paths

    def _parse_file(self, file_path: Path) -> str:
        """
        PDF/Excel 파일 파싱하여 텍스트 반환.

        Args:
            file_path: 파일 경로

        Returns:
            추출된 텍스트 내용
        """
        suffix = file_path.suffix.lower()

        if suffix == ".pdf":
            parser = self._get_pdf_parser()
            result = parser.extract_all(file_path)
            text_parts = [result.get("text", "")]

            # 테이블 데이터도 텍스트로 변환
            for table in result.get("tables", []):
                try:
                    import pandas as pd
                    if isinstance(table, pd.DataFrame):
                        text_parts.append(table.to_string())
                    else:
                        for row in table:
                            text_parts.append(" | ".join(str(c) for c in row if c))
                except Exception:
                    if isinstance(table, list):
                        for row in table:
                            text_parts.append(
                                " | ".join(str(c) for c in row if c)
                            )

            return "\n".join(text_parts)

        elif suffix in (".xlsx", ".xls"):
            parser = self._get_excel_parser()
            sheets = parser.parse(file_path)
            text_parts = []

            for sheet_name, data in sheets.items():
                text_parts.append(f"[Sheet: {sheet_name}]")
                try:
                    import pandas as pd
                    if isinstance(data, pd.DataFrame):
                        text_parts.append(data.to_string())
                    else:
                        for row in data:
                            text_parts.append(
                                " | ".join(str(c) for c in row if c is not None)
                            )
                except Exception:
                    if isinstance(data, list):
                        for row in data:
                            text_parts.append(
                                " | ".join(str(c) for c in row if c is not None)
                            )

            return "\n".join(text_parts)

        return ""

    def _extract_quotes(self, text: str, vendor: str) -> list:
        """
        규칙 기반 금액 추출.

        Args:
            text: 파싱된 텍스트
            vendor: 업체명

        Returns:
            QuoteOption 리스트
        """
        import re

        extractor = self._get_quote_extractor()
        options = extractor.extract_from_text(text)

        # $ 접두사 금액을 원본 텍스트에서 추출하여 USD 보정
        dollar_amounts = set()
        for m in re.finditer(r'\$\s*([\d,]+(?:\.\d+)?)', text):
            try:
                amt = float(m.group(1).replace(",", ""))
                dollar_amounts.add(amt)
            except ValueError:
                pass

        for opt in options:
            # currency 보정: 원본에 $ 접두사가 있으면 USD
            if opt.total_amount is not None:
                opt_amt = float(opt.total_amount)
                if opt_amt in dollar_amounts:
                    opt.currency = "USD"

            # source_file에 vendor 정보 추가
            if not opt.source_file:
                opt.source_file = f"attachment:{vendor}"

        return options

    def _extract_quotes_from_excel(self, file_path: Path) -> list:
        """
        Excel 시트에서 견적 금액을 직접 추출.

        총합/정가/금액 키워드가 있는 행에서 숫자 셀을 찾아 추출.
        통화는 금액 범위로 추론 (1000~1M = USD, 1M+ = KRW).

        Args:
            file_path: Excel 파일 경로

        Returns:
            QuoteOption 리스트
        """
        try:
            from openpyxl import load_workbook
        except ImportError:
            raise ImportError("openpyxl not available: pip install openpyxl")

        # Import QuoteOption
        try:
            from models_v2 import QuoteOption
        except ImportError:
            try:
                from .models_v2 import QuoteOption
            except ImportError:
                return []

        options = []
        wb = load_workbook(file_path, data_only=True)

        # 키워드 패턴
        quote_keywords = ["총합", "정가", "금액", "할인", "이용", "annual", "total", "pricing"]
        discount_keywords = ["할인", "discount", "적용"]

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]

            # 시트별 옵션명 추출
            sheet_option_name = sheet_name

            # 시트별로 후보 행 수집 (할인가 vs 정가)
            discount_candidates = []
            regular_candidates = []

            # 각 행을 순회하며 키워드 + 숫자 셀 찾기
            for row_idx, row in enumerate(ws.iter_rows(min_row=1, values_only=True), start=1):
                if not row:
                    continue

                # 행 내용을 문자열로 변환
                row_text = " ".join(str(cell).lower() for cell in row if cell is not None)

                # 키워드가 있는지 확인
                has_quote_keyword = any(kw in row_text for kw in quote_keywords)
                if not has_quote_keyword:
                    continue

                # 행에서 숫자 셀 추출 (float/int)
                amounts = []
                for cell_idx, cell in enumerate(row):
                    if isinstance(cell, (int, float)):
                        # 1000 미만은 단가로 간주하여 무시
                        if cell >= 1000:
                            amounts.append(cell)

                if not amounts:
                    continue

                # 할인가 우선 사용 (할인 키워드가 있고 금액이 1개 이상)
                is_discount_row = any(kw in row_text for kw in discount_keywords)

                # 한 행당 하나의 금액만 추출 (가장 큰 값 = 총합)
                amount = max(amounts)

                if is_discount_row:
                    discount_candidates.append((amount, row_idx))
                else:
                    regular_candidates.append((amount, row_idx))

            # 시트당 1개 옵션만 추출 (할인가 우선, 없으면 정가)
            if discount_candidates:
                # 할인가 중 가장 큰 금액 사용
                amount, row_idx = max(discount_candidates, key=lambda x: x[0])
                currency = "USD" if 1_000 <= amount < 1_000_000 else "KRW"
                option_name = f"{sheet_option_name} (할인가)"
            elif regular_candidates:
                # 정가 중 가장 큰 금액 사용
                amount, row_idx = max(regular_candidates, key=lambda x: x[0])
                currency = "USD" if 1_000 <= amount < 1_000_000 else "KRW"
                option_name = f"{sheet_option_name}"
            else:
                # 해당 시트에서 금액 없음
                continue

            # QuoteOption 생성
            option = QuoteOption(
                option_id=f"{sheet_name}_{row_idx}",
                option_name=option_name[:50],
                total_amount=amount,
                currency=currency,
                source_file=f"{file_path.name} - {sheet_name}",
                extracted_at=datetime.now(),
                extraction_method="excel_direct",
                confidence=0.9,
            )
            options.append(option)

        return options

    def _generate_ai_summary(
        self,
        vendor: str,
        parsed_content: str,
        extracted_quotes: list,
    ) -> str:
        """
        파싱된 내용으로 견적 요약 생성.

        QuoteFormatter를 사용하여 업체별 상황에 맞는 간결한 포맷을 적용합니다.
        QuoteFormatter 사용 불가 시 기존 인라인 로직으로 fallback합니다.

        Args:
            vendor: 업체명
            parsed_content: 파싱된 전체 텍스트
            extracted_quotes: QuoteOption 리스트

        Returns:
            요약 문자열 (예: "$42.5K~$388K (4옵션)")
        """
        if not extracted_quotes:
            if parsed_content.strip():
                return "첨부파일 분석 완료 (금액 미감지)"
            return "N/A"

        try:
            if QuoteFormatter is not None:
                formatter = QuoteFormatter()
                return formatter.format_quote(
                    vendor_name=vendor,
                    options=extracted_quotes,
                )
        except Exception:
            pass

        # Fallback: 기존 인라인 로직
        amounts = []
        for opt in extracted_quotes:
            if opt.total_amount is not None:
                amounts.append((float(opt.total_amount), opt.currency))
        if not amounts:
            return "금액 미감지"
        amounts.sort(key=lambda x: x[0])
        if len(amounts) == 1:
            return self._format_amount(amounts[0][0], amounts[0][1])
        else:
            min_amt = self._format_amount(amounts[0][0], amounts[0][1])
            max_amt = self._format_amount(amounts[-1][0], amounts[-1][1])
            return f"{min_amt} ~ {max_amt} ({len(amounts)}개 옵션)"

    @staticmethod
    def _format_amount(amount, currency: str = "KRW") -> str:
        """금액을 읽기 쉬운 문자열로 변환. (Deprecated: QuoteFormatter 사용 권장)"""
        amt = float(amount)

        if currency == "USD":
            # USD 전용 포맷
            if amt >= 1_000_000:
                return f"${amt / 1_000_000:.1f}M"
            elif amt >= 1_000:
                return f"${amt / 1_000:.1f}K"
            return f"${amt:,.0f}"
        else:
            # KRW
            if amt >= 100_000_000:
                return f"{amt / 100_000_000:.1f}억원"
            elif amt >= 10_000:
                return f"{amt / 10_000:,.0f}만원"
            return f"{amt:,.0f}원"


class ListsSyncManager:
    """Manage synchronization with Slack Lists."""

    def __init__(self):
        """Initialize with both bot and user clients."""
        self.bot_client = SlackClient()
        try:
            self.user_client = SlackUserClient()
        except Exception as e:
            print(f"Warning: User token not available, Lists API disabled: {e}")
            self.user_client = None

    def get_list_items(self) -> list[dict]:
        """
        Fetch all items from Slack List.

        Returns:
            List of item dictionaries with parsed fields
        """
        if not self.user_client:
            raise RuntimeError("User token required for Lists API")

        result = self.user_client.get_list_items(LISTS_CONFIG["list_id"])

        if not result.get("ok"):
            raise RuntimeError(f"Failed to fetch list items: {result}")

        items = []
        for item in result.get("items", []):
            parsed = self._parse_item(item)
            if parsed:
                items.append(parsed)

        return items

    def _parse_item(self, item: dict) -> Optional[dict]:
        """Parse a list item into structured data."""
        fields_list = item.get("fields", [])
        row_id = item.get("id", "")

        # Convert list of fields to dict by column_id
        fields = {}
        for field in fields_list:
            col_id = field.get("column_id", "")
            if col_id:
                fields[col_id] = field

        def get_text_field(column_id: str) -> str:
            """Extract text from rich_text field."""
            field = fields.get(column_id, {})
            # First try 'text' directly
            if field.get("text"):
                return field.get("text", "")
            # Then try rich_text parsing
            rich_text = field.get("rich_text", [])
            if rich_text:
                elements = rich_text[0].get("elements", [])
                if elements:
                    sections = elements[0].get("elements", [])
                    if sections:
                        return sections[0].get("text", "")
            return ""

        def get_select_field(column_id: str) -> str:
            """Extract selected option from select field."""
            field = fields.get(column_id, {})
            # Try 'select' array first
            select_arr = field.get("select", [])
            if select_arr:
                option_id = select_arr[0]
            else:
                option_id = field.get("value", "")
            # Reverse lookup option name
            for name, opt_id in LISTS_CONFIG["status_options"].items():
                if opt_id == option_id:
                    return name
            return ""

        def get_date_field(column_id: str) -> Optional[datetime]:
            """Extract date from date field."""
            field = fields.get(column_id, {})
            # Try 'date' array first
            date_arr = field.get("date", [])
            if date_arr:
                date_str = date_arr[0]
            else:
                date_str = field.get("value", "")
            if date_str:
                try:
                    return datetime.strptime(date_str, "%Y-%m-%d")
                except ValueError:
                    pass
            return None

        return {
            "row_id": row_id,
            "vendor_name": get_text_field(LISTS_CONFIG["columns"]["vendor_name"]),
            "status": get_select_field(LISTS_CONFIG["columns"]["status"]),
            "quote": get_text_field(LISTS_CONFIG["columns"]["quote"]),
            "last_contact": get_date_field(LISTS_CONFIG["columns"]["last_contact"]),
            "next_action": get_text_field(LISTS_CONFIG["columns"]["next_action"]),
            "notes": get_text_field(LISTS_CONFIG["columns"]["notes"]),
        }

    def update_item(
        self,
        vendor_name: str,
        status: Optional[str] = None,
        quote: Optional[str] = None,
        last_contact: Optional[datetime] = None,
        next_action: Optional[str] = None,
        notes: Optional[str] = None,
    ) -> bool:
        """
        Update a vendor item in Slack List.

        Args:
            vendor_name: Name of vendor to update
            status: New status (must match status_options keys)
            quote: Quote text
            last_contact: Last contact date
            next_action: Next action text
            notes: Notes text

        Returns:
            True if successful
        """
        if not self.user_client:
            raise RuntimeError("User token required for Lists API")

        row_id = LISTS_CONFIG["items"].get(vendor_name)
        if not row_id:
            raise ValueError(f"Unknown vendor: {vendor_name}")

        cells = []

        if status:
            option_id = LISTS_CONFIG["status_options"].get(status)
            if option_id:
                cells.append({
                    "column_id": LISTS_CONFIG["columns"]["status"],
                    "row_id": row_id,
                    "select": [option_id],  # select는 배열 형식
                })

        if quote:
            cells.append({
                "column_id": LISTS_CONFIG["columns"]["quote"],
                "row_id": row_id,
                "rich_text": [{"type": "rich_text", "elements": [{"type": "rich_text_section", "elements": [{"type": "text", "text": quote}]}]}],
            })

        if last_contact:
            cells.append({
                "column_id": LISTS_CONFIG["columns"]["last_contact"],
                "row_id": row_id,
                "date": [last_contact.strftime("%Y-%m-%d")],  # date도 배열 형식
            })

        if next_action:
            cells.append({
                "column_id": LISTS_CONFIG["columns"]["next_action"],
                "row_id": row_id,
                "rich_text": [{"type": "rich_text", "elements": [{"type": "rich_text_section", "elements": [{"type": "text", "text": next_action}]}]}],
            })

        if notes:
            cells.append({
                "column_id": LISTS_CONFIG["columns"]["notes"],
                "row_id": row_id,
                "rich_text": [{"type": "rich_text", "elements": [{"type": "rich_text_section", "elements": [{"type": "text", "text": notes}]}]}],
            })

        if not cells:
            return False

        result = self.user_client._client.api_call(
            "slackLists.items.update",
            json={
                "list_id": LISTS_CONFIG["list_id"],
                "cells": cells,
            },
        )

        return result.data.get("ok", False)

    def generate_summary_message(self, items: list[dict]) -> str:
        """
        Generate summary message for Slack posting.

        Args:
            items: List of vendor items

        Returns:
            Formatted Slack message with blocks support
        """
        now = datetime.now().strftime("%Y-%m-%d %H:%M")
        team_id = LISTS_CONFIG["team_id"]
        list_id = LISTS_CONFIG["list_id"]
        list_url = f"https://ggproduction.slack.com/lists/{team_id}/{list_id}"

        lines = [
            f"<{list_url}|*WSOPTV Vendor Status*> (auto-update: {now})",
            "",
        ]

        # Group by status
        status_groups = {}
        for item in items:
            status = item.get("status", "미정")
            if status not in status_groups:
                status_groups[status] = []
            status_groups[status].append(item)

        # Status emoji mapping
        status_emoji = {
            "협상 중": ":large_green_circle:",
            "검토 중": ":large_yellow_circle:",
            "견적 대기": ":hourglass:",
            "보류": ":red_circle:",
        }

        for status, vendors in status_groups.items():
            emoji = status_emoji.get(status, "📋")
            lines.append(f"{emoji} *{status}* ({len(vendors)}개)")

            for vendor in vendors:
                name = vendor.get("vendor_name", "")
                quote = vendor.get("quote", "-")
                next_act = vendor.get("next_action", "-")
                lines.append(f"  • {name}: {quote}")
                if next_act and next_act != "-":
                    lines.append(f"    └ 다음: {next_act}")

            lines.append("")

        lines.append("_상세 현황: VENDOR-DASHBOARD.md 참조_")

        return "\n".join(lines)

    def post_summary(self, message: str, update_existing: bool = True) -> Optional[str]:
        """
        Post or update summary message to project channel.

        Args:
            message: Message text
            update_existing: If True, update pinned message instead of posting new

        Returns:
            Message timestamp if successful
        """
        channel_id = LISTS_CONFIG["channel_id"]
        pinned_ts = LISTS_CONFIG.get("pinned_message_ts")

        if update_existing and pinned_ts:
            # Update existing pinned message
            result = self.bot_client._client.chat_update(
                channel=channel_id,
                ts=pinned_ts,
                text=message,
            )
            return result.data.get("ts") if result.data.get("ok") else None
        else:
            # Post new message (fallback)
            result = self.bot_client.send_message(channel_id, message)
            return result.ts if result and result.ok else None


def sync_lists_to_dashboard(
    dashboard_path: Path,
    dry_run: bool = False,
) -> SyncResult:
    """
    Sync Slack Lists to VENDOR-DASHBOARD.md.

    Args:
        dashboard_path: Path to VENDOR-DASHBOARD.md
        dry_run: If True, only show what would be updated

    Returns:
        SyncResult with counts
    """
    result = SyncResult()

    try:
        manager = ListsSyncManager()
        items = manager.get_list_items()

        print(f"Fetched {len(items)} items from Slack List")

        for item in items:
            print(f"  - {item['vendor_name']}: {item['status']}")
            result.added += 1

        if dry_run:
            print("\n=== DRY RUN ===")
            print("Would update VENDOR-DASHBOARD.md with above data")
            return result

        # TODO: Parse and update VENDOR-DASHBOARD.md
        # For now, just output the data
        print("\nLists data fetched successfully")

    except Exception as e:
        print(f"Error: {e}")
        result.errors += 1

    return result


def post_daily_summary(dry_run: bool = False) -> bool:
    """
    Post daily summary to Slack channel.

    Args:
        dry_run: If True, only show message without posting

    Returns:
        True if successful
    """
    try:
        manager = ListsSyncManager()
        items = manager.get_list_items()

        message = manager.generate_summary_message(items)

        if dry_run:
            print("\n=== DRY RUN - Message Preview ===")
            print(message)
            return True

        ts = manager.post_summary(message)
        if ts:
            print(f"Summary posted to Slack (ts: {ts})")
            return True
        else:
            print("Failed to post summary")
            return False

    except Exception as e:
        print(f"Error: {e}")
        return False


class IncrementalSyncState:
    """Manage incremental sync state for Slack Lists updates."""

    STATE_FILE = Path("C:/claude/wsoptv_ott/docs/management/.slacklist_state.json")

    def __init__(self):
        """Load existing state or initialize new."""
        self.state = self._load_state()

    def _load_state(self) -> dict:
        """Load state from file or return default."""
        if self.STATE_FILE.exists():
            try:
                return json.loads(self.STATE_FILE.read_text(encoding="utf-8"))
            except Exception:
                pass
        return self._default_state()

    def _default_state(self) -> dict:
        """Return default state structure."""
        return {
            "version": "1.0",
            "last_sync": None,
            "processed": {
                "gmail": {"last_email_date": None, "email_ids": []},
                "slack": {"last_ts": None, "message_ts": []},
            },
            "vendors": {},
            "pending_changes": [],
        }

    def save(self):
        """Save current state to file."""
        self.state["last_sync"] = datetime.now().isoformat()
        self.STATE_FILE.write_text(
            json.dumps(self.state, indent=2, ensure_ascii=False, default=str),
            encoding="utf-8",
        )

    def is_gmail_processed(self, email_id: str) -> bool:
        """Check if email was already processed."""
        return email_id in self.state["processed"]["gmail"]["email_ids"]

    def mark_gmail_processed(self, email_id: str):
        """Mark email as processed."""
        if email_id not in self.state["processed"]["gmail"]["email_ids"]:
            self.state["processed"]["gmail"]["email_ids"].append(email_id)

    def is_slack_processed(self, ts: str) -> bool:
        """Check if Slack message was already processed."""
        return ts in self.state["processed"]["slack"]["message_ts"]

    def mark_slack_processed(self, ts: str):
        """Mark Slack message as processed."""
        if ts not in self.state["processed"]["slack"]["message_ts"]:
            self.state["processed"]["slack"]["message_ts"].append(ts)

    def get_last_slack_ts(self) -> Optional[str]:
        """Get last processed Slack timestamp."""
        return self.state["processed"]["slack"]["last_ts"]

    def update_last_slack_ts(self, ts: str):
        """Update last processed Slack timestamp."""
        self.state["processed"]["slack"]["last_ts"] = ts

    def add_pending_change(self, change: dict):
        """Add a pending change for user review."""
        change["detected_at"] = datetime.now().isoformat()
        self.state["pending_changes"].append(change)

    def get_pending_changes(self) -> list:
        """Get all pending changes."""
        return self.state["pending_changes"]

    def clear_pending_changes(self):
        """Clear pending changes after applying."""
        self.state["pending_changes"] = []

    def update_vendor(self, vendor_name: str, updates: dict):
        """Update vendor state."""
        if vendor_name not in self.state["vendors"]:
            self.state["vendors"][vendor_name] = {"changes_history": []}

        vendor = self.state["vendors"][vendor_name]
        for key, value in updates.items():
            if key != "changes_history":
                old_value = vendor.get(key)
                if old_value != value:
                    vendor["changes_history"].append({
                        "field": key,
                        "old": old_value,
                        "new": value,
                        "changed_at": datetime.now().isoformat(),
                    })
                vendor[key] = value

    def get_vendor_state(self, vendor_name: str) -> Optional[dict]:
        """Get current vendor state."""
        return self.state["vendors"].get(vendor_name)

    def get_stats(self) -> dict:
        """Get sync statistics."""
        return {
            "last_sync": self.state.get("last_sync"),
            "gmail_processed": len(self.state["processed"]["gmail"]["email_ids"]),
            "slack_processed": len(self.state["processed"]["slack"]["message_ts"]),
            "vendors_tracked": len(self.state["vendors"]),
            "pending_changes": len(self.state["pending_changes"]),
        }


class VendorChangeDetector:
    """Detect vendor-related changes from Gmail and Slack messages."""

    # Vendor detection patterns (domain → vendor name)
    VENDOR_DOMAINS = {
        "brightcove": "Brightcove",
        "megazone": "메가존클라우드",
        "megazonecloud": "메가존클라우드",
        "mz.co.kr": "메가존클라우드",
        "vimeo": "Vimeo OTT",
        "jrmax": "Vimeo OTT",  # 한국 대리점
        "wecandeo": "맑음소프트 (WECANDEO)",
        "malgum": "맑음소프트 (WECANDEO)",
    }

    # Quote detection patterns for subject lines
    QUOTE_SUBJECT_PATTERNS = [
        "pricing proposal",
        "price quote",
        "quotation",
        "견적",
        "견적서",
        "견적 전달",
        "견적 제안",
        "proposal",
    ]

    # Status transition keywords - categorized by direction
    # POSITIVE: Status should improve or stay active
    POSITIVE_KEYWORDS = {
        # 한글 견적 패턴
        "견적 수령": ("견적 대기", "검토 중"),
        "견적 도착": ("견적 대기", "검토 중"),
        "견적서 첨부": ("견적 대기", "검토 중"),
        "견적 전달": ("견적 대기", "검토 중"),
        "견적 제안": ("견적 대기", "검토 중"),
        "견적 안내": ("견적 대기", "검토 중"),
        "견적을": ("견적 대기", "검토 중"),  # "견적을 전달", "견적을 수정" 등
        # 영문 견적 패턴
        "pricing proposal": ("견적 대기", "검토 중"),
        "price quote": ("견적 대기", "검토 중"),
        "quotation": ("견적 대기", "검토 중"),
        # 수정 견적 (협상 단계)
        "수정 견적": ("검토 중", "협상 중"),
        "revised quote": ("검토 중", "협상 중"),
        "updated pricing": ("검토 중", "협상 중"),
        # 기타 협상 패턴
        "협상 시작": ("검토 중", "협상 중"),
        "조건 협의": ("검토 중", "협상 중"),
        "계약서 검토": ("협상 중", "계약 진행"),
        "법무 검토": ("협상 중", "계약 진행"),
        "미팅 예정": ("*", "협상 중"),
        "회의 예정": ("*", "협상 중"),
    }

    # NEGATIVE: Status should decline (requires explicit approval)
    NEGATIVE_KEYWORDS = {
        "프로젝트 취소": ("*", "제외"),
        "진행 불가": ("*", "제외"),
        "서비스 종료": ("*", "제외"),
        "매각 진행": ("*", "제외"),  # More specific than just "매각"
        "협력 중단": ("*", "제외"),
    }

    # NEUTRAL: Just informational, no automatic status change
    NEUTRAL_KEYWORDS = ["미수령", "무응답", "대기 중", "검토 중", "확인 중"]

    # Legacy combined dict for backward compatibility
    STATUS_KEYWORDS = {**POSITIVE_KEYWORDS, **NEGATIVE_KEYWORDS}

    # Action keywords for next_action field
    ACTION_KEYWORDS = [
        "미팅", "회의", "연락", "확인", "검토", "요청", "피드백", "리마인더", "follow-up"
    ]

    def __init__(self):
        """Initialize detector with pattern cache."""
        self._changes = []

    def detect_vendor_from_email(self, sender: str, subject: str) -> Optional[str]:
        """Detect vendor from email sender or subject."""
        email_domain = sender.split("@")[-1].lower() if "@" in sender else ""

        for domain_key, vendor_name in self.VENDOR_DOMAINS.items():
            if domain_key in email_domain or domain_key in subject.lower():
                return vendor_name
        return None

    def detect_status_change(
        self,
        text: str,
        is_vendor_email: bool = False,
    ) -> Optional[tuple]:
        """
        Detect status change from text content.

        Args:
            text: Message text to analyze
            is_vendor_email: True if this is an email FROM the vendor (not TO)

        Returns:
            Tuple of (from_status, to_status, keyword, direction) or None
            direction: "positive", "negative", or None
        """
        text_lower = text.lower()

        # Check positive keywords first (higher priority)
        for keyword, (from_status, to_status) in self.POSITIVE_KEYWORDS.items():
            if keyword in text_lower:
                return (from_status, to_status, keyword, "positive")

        # Check negative keywords (requires explicit approval)
        for keyword, (from_status, to_status) in self.NEGATIVE_KEYWORDS.items():
            if keyword in text_lower:
                return (from_status, to_status, keyword, "negative")

        # If vendor sent us an email, that's a positive signal (active communication)
        # Don't suggest status decline just because neutral keywords exist
        if is_vendor_email:
            # Vendor email = active communication, no automatic status decline
            return None

        return None

    def detect_quote_signal(
        self,
        subject: str,
        body: str,
        has_attachments: bool = False,
        is_vendor_email: bool = False,
    ) -> bool:
        """
        Detect quote/pricing signals from email.

        Args:
            subject: Email subject line
            body: Email body text
            has_attachments: Whether email has attachments
            is_vendor_email: Whether email is from vendor

        Returns:
            True if quote is detected (2+ indicators present)
        """
        indicators = 0

        # Indicator 1: Subject contains quote keywords
        subject_lower = subject.lower()
        if any(pattern in subject_lower for pattern in self.QUOTE_SUBJECT_PATTERNS):
            indicators += 1

        # Indicator 2: Has attachments (likely quote file)
        if has_attachments:
            indicators += 1

        # Indicator 3: From vendor domain
        if is_vendor_email:
            indicators += 1

        # Indicator 4: Body contains quote keywords
        body_lower = body.lower()
        if any(pattern in body_lower for pattern in self.QUOTE_SUBJECT_PATTERNS):
            indicators += 1

        # Require 2+ indicators to avoid false positives
        return indicators >= 2

    def detect_action_item(self, text: str) -> Optional[str]:
        """Detect action item from text content."""
        import re

        # Pattern: date-like content with action keyword
        date_pattern = r"(\d{1,2}[-/]\d{1,2}|\d{2}-\d{2})"
        action_pattern = "|".join(self.ACTION_KEYWORDS)

        if re.search(date_pattern, text) and re.search(action_pattern, text, re.IGNORECASE):
            # Extract the action item portion
            lines = text.split("\n")
            for line in lines:
                if re.search(date_pattern, line) and re.search(action_pattern, line, re.IGNORECASE):
                    return line.strip()[:100]  # Limit to 100 chars
        return None

    def analyze_gmail_messages(self, emails: list[dict]) -> list[dict]:
        """
        Analyze Gmail messages for vendor changes.

        Args:
            emails: List of email dicts with keys: sender, subject, body, date, has_attachments

        Returns:
            List of detected changes
        """
        changes = []

        for email in emails:
            sender = email.get("sender", "")
            subject = email.get("subject", "")
            body = email.get("body", "")
            has_attachments = email.get("has_attachments", False)

            vendor = self.detect_vendor_from_email(sender, subject)
            if not vendor:
                continue

            # Check if email is FROM the vendor (not TO the vendor)
            # Vendor email = positive signal (active communication)
            sender_domain = sender.split("@")[-1].lower() if "@" in sender else ""
            is_vendor_email = any(
                domain_key in sender_domain
                for domain_key in self.VENDOR_DOMAINS.keys()
            )

            change = {
                "source": "gmail",
                "vendor": vendor,
                "date": email.get("date"),
                "subject": subject,
                "is_vendor_email": is_vendor_email,
            }

            # Enhanced quote detection
            quote_detected = self.detect_quote_signal(
                subject=subject,
                body=body,
                has_attachments=has_attachments,
                is_vendor_email=is_vendor_email,
            )

            if quote_detected:
                change["quote_detected"] = True
                change["status_change"] = {
                    "from": "견적 대기",
                    "to": "검토 중",
                    "trigger": "quote_signal",
                    "direction": "positive",
                }

            # Check for status change (with vendor email context)
            full_text = f"{subject} {body}"
            status_change = self.detect_status_change(full_text, is_vendor_email=is_vendor_email)
            if status_change and not quote_detected:
                # Only apply text-based status change if quote signal didn't already detect it
                change["status_change"] = {
                    "from": status_change[0],
                    "to": status_change[1],
                    "trigger": status_change[2],
                    "direction": status_change[3],  # positive or negative
                }

            # Check for action item
            action = self.detect_action_item(full_text)
            if action:
                change["next_action"] = action

            # Always update last_contact for vendor emails
            change["last_contact"] = email.get("date")

            changes.append(change)

        return changes

    def analyze_slack_messages(self, messages: list) -> list[dict]:
        """
        Analyze Slack messages for vendor-related changes.

        Args:
            messages: List of SlackMessage objects or dicts with keys: text, user, ts

        Returns:
            List of detected changes
        """
        changes = []

        for msg in messages:
            # Handle both pydantic models and dicts
            if hasattr(msg, "text"):
                text = getattr(msg, "text", "")
            else:
                text = msg.get("text", "") if isinstance(msg, dict) else ""

            # Detect vendor mentions in message
            detected_vendor = None
            for domain_key, vendor_name in self.VENDOR_DOMAINS.items():
                if domain_key in text.lower() or vendor_name.lower() in text.lower():
                    detected_vendor = vendor_name
                    break

            if not detected_vendor:
                # Also check for vendor names in Korean
                vendor_keywords = {
                    "메가존": "메가존클라우드",
                    "브라이트코브": "Brightcove",
                    "비메오": "Vimeo OTT",
                    "맑음소프트": "맑음소프트 (WECANDEO)",
                }
                for keyword, vendor_name in vendor_keywords.items():
                    if keyword in text:
                        detected_vendor = vendor_name
                        break

            if not detected_vendor:
                continue

            # Get ts from model or dict
            msg_ts = getattr(msg, "ts", None) if hasattr(msg, "ts") else (msg.get("ts") if isinstance(msg, dict) else None)

            change = {
                "source": "slack",
                "vendor": detected_vendor,
                "ts": msg_ts,
                "text_preview": text[:100],
            }

            # Check for status change (Slack messages are internal, not vendor emails)
            status_change = self.detect_status_change(text, is_vendor_email=False)
            if status_change:
                change["status_change"] = {
                    "from": status_change[0],
                    "to": status_change[1],
                    "trigger": status_change[2],
                    "direction": status_change[3],
                }

            # Check for action item
            action = self.detect_action_item(text)
            if action:
                change["next_action"] = action

            changes.append(change)

        return changes

    def merge_changes(self, gmail_changes: list[dict], slack_changes: list[dict]) -> dict[str, dict]:
        """
        Merge changes from Gmail and Slack by vendor.

        Args:
            gmail_changes: Changes detected from Gmail
            slack_changes: Changes detected from Slack

        Returns:
            Dict mapping vendor name to aggregated changes
        """
        merged = {}

        for change in gmail_changes + slack_changes:
            vendor = change.get("vendor")
            if not vendor:
                continue

            if vendor not in merged:
                merged[vendor] = {
                    "vendor": vendor,
                    "sources": [],
                    "status_changes": [],
                    "next_actions": [],
                    "last_contact": None,
                    "quote_count": 0,
                    "quote_emails": [],
                }

            merged[vendor]["sources"].append(change.get("source"))

            if "status_change" in change:
                merged[vendor]["status_changes"].append(change["status_change"])

            if "next_action" in change:
                merged[vendor]["next_actions"].append(change["next_action"])

            if "last_contact" in change and change["last_contact"]:
                if not merged[vendor]["last_contact"] or change["last_contact"] > merged[vendor]["last_contact"]:
                    merged[vendor]["last_contact"] = change["last_contact"]

            # 견적 감지 카운트
            if change.get("quote_detected"):
                merged[vendor]["quote_count"] += 1
                merged[vendor]["quote_emails"].append({
                    "date": change.get("date"),
                    "subject": change.get("subject"),
                })

        return merged


def intelligent_update_slacklist(
    days: int = 7,
    dry_run: bool = False,
    auto_approve: bool = False,
    vendor_filter: Optional[str] = None,
    incremental: bool = True,
) -> dict:
    """
    Intelligently update Slack Lists based on Gmail and Slack message analysis.

    This is the main entry point for the "/auto update slacklist" workflow.

    Args:
        days: Number of days to analyze (ignored if incremental=True and state exists)
        dry_run: If True, show changes without applying
        auto_approve: If True, apply changes without confirmation
        vendor_filter: Filter to specific vendor
        incremental: If True, only process new messages since last sync

    Returns:
        Dict with update results
    """
    import sys
    sys.path.insert(0, str(Path(__file__).parents[2]))

    from lib.gmail import GmailClient
    from lib.slack import SlackClient

    # Load incremental state
    sync_state = IncrementalSyncState()
    stats = sync_state.get_stats()

    print(f"\n{'='*60}")
    print("[SYNC] Intelligent Slack Lists Update")
    print(f"{'='*60}")

    if incremental and stats["last_sync"]:
        print(f"Mode: INCREMENTAL (since {stats['last_sync'][:19]})")
        print(f"  Previously processed: Gmail {stats['gmail_processed']}, Slack {stats['slack_processed']}")
    else:
        print(f"Mode: FULL SCAN ({days} days)")

    if vendor_filter:
        print(f"Vendor Filter: {vendor_filter}")

    results = {
        "analyzed": {"gmail": 0, "slack": 0},
        "changes_detected": [],
        "changes_applied": [],
        "errors": [],
    }

    # Step 1: Collect data
    print("\n[Step 1] Collecting data...")

    # Gmail - search with multiple strategies to find vendor emails
    gmail_client = None
    try:
        gmail_client = GmailClient()
        from datetime import datetime, timedelta
        after_date = (datetime.now() - timedelta(days=days)).strftime("%Y/%m/%d")

        # Strategy 1: Label-based search (if label exists)
        labeled_emails = gmail_client.list_emails(
            query=f"label:wsoptv after:{after_date}",
            max_results=100,
        )

        # Strategy 2: Keyword-based search (catches unlabeled emails)
        keyword_queries = [
            f"(WSOPTV OR WSOP) after:{after_date}",
            f"(from:brightcove OR from:megazone OR from:vimeo) after:{after_date}",
            f"(RFP OR OTT OR 스트리밍) after:{after_date}",
        ]

        keyword_emails = []
        seen_ids = {getattr(e, "id", None) or getattr(e, "message_id", "") for e in labeled_emails}

        for kw_query in keyword_queries:
            try:
                kw_results = gmail_client.list_emails(query=kw_query, max_results=50)
                for email in kw_results:
                    email_id = getattr(email, "id", None) or getattr(email, "message_id", "")
                    if email_id and email_id not in seen_ids:
                        keyword_emails.append(email)
                        seen_ids.add(email_id)
            except Exception:
                pass  # Continue with other queries if one fails

        all_emails = list(labeled_emails) + keyword_emails
        if keyword_emails:
            print(f"  [INFO] Found {len(keyword_emails)} unlabeled emails via keyword search")

        # Filter out already processed emails in incremental mode
        if incremental:
            emails = []
            for email in all_emails:
                email_id = getattr(email, "id", None) or getattr(email, "message_id", "")
                if not sync_state.is_gmail_processed(email_id):
                    emails.append(email)
            print(f"  [OK] Gmail: {len(emails)} new / {len(all_emails)} total")
        else:
            emails = all_emails
            print(f"  [OK] Gmail: {len(emails)} emails")

        results["analyzed"]["gmail"] = len(emails)
    except Exception as e:
        print(f"  [ERR] Gmail: {e}")
        results["errors"].append(f"Gmail: {e}")
        emails = []

    # Slack
    try:
        slack_client = SlackClient()
        all_slack_messages = slack_client.get_history(
            channel=LISTS_CONFIG["channel_id"],
            limit=200,
        )

        # Filter out already processed messages in incremental mode
        if incremental:
            slack_messages = []
            for msg in all_slack_messages:
                msg_ts = getattr(msg, "ts", None) if hasattr(msg, "ts") else msg.get("ts") if isinstance(msg, dict) else None
                if msg_ts and not sync_state.is_slack_processed(msg_ts):
                    slack_messages.append(msg)
            print(f"  [OK] Slack: {len(slack_messages)} new / {len(all_slack_messages)} total")
        else:
            slack_messages = all_slack_messages
            print(f"  [OK] Slack: {len(slack_messages)} messages")

        results["analyzed"]["slack"] = len(slack_messages)
    except Exception as e:
        print(f"  [ERR] Slack: {e}")
        results["errors"].append(f"Slack: {e}")
        slack_messages = []

    # Current Slack Lists
    try:
        manager = ListsSyncManager()
        current_items = manager.get_list_items()
        print(f"  [OK] Slack Lists: {len(current_items)} vendors")
    except Exception as e:
        print(f"  [ERR] Slack Lists: {e}")
        results["errors"].append(f"Slack Lists: {e}")
        current_items = []

    # Step 2: Analyze changes
    print("\n[Step 2] Analyzing changes...")

    detector = VendorChangeDetector()

    # Convert Gmail to expected format (GmailMessage is pydantic model)
    gmail_formatted = []
    for email in emails:
        email_id = getattr(email, "id", None) or getattr(email, "message_id", "")
        # Check for attachments
        has_attachments = False
        if hasattr(email, "attachments"):
            has_attachments = bool(getattr(email, "attachments", []))
        elif hasattr(email, "has_attachments"):
            has_attachments = getattr(email, "has_attachments", False)

        gmail_formatted.append({
            "id": email_id,
            "sender": getattr(email, "from_", "") or getattr(email, "sender", ""),
            "subject": getattr(email, "subject", ""),
            "body": getattr(email, "snippet", ""),
            "date": getattr(email, "date", ""),
            "has_attachments": has_attachments,
        })

    gmail_changes = detector.analyze_gmail_messages(gmail_formatted)
    slack_changes = detector.analyze_slack_messages(slack_messages)

    print(f"  Gmail changes: {len(gmail_changes)}")
    print(f"  Slack changes: {len(slack_changes)}")

    # Mark processed (will save later if not dry_run)
    if incremental:
        for email in emails:
            email_id = getattr(email, "id", None) or getattr(email, "message_id", "")
            if email_id:
                sync_state.mark_gmail_processed(email_id)
        for msg in slack_messages:
            msg_ts = getattr(msg, "ts", None) if hasattr(msg, "ts") else msg.get("ts") if isinstance(msg, dict) else None
            if msg_ts:
                sync_state.mark_slack_processed(msg_ts)

    merged = detector.merge_changes(gmail_changes, slack_changes)

    # Apply vendor filter if specified
    if vendor_filter:
        filter_lower = vendor_filter.lower()
        merged = {k: v for k, v in merged.items() if filter_lower in k.lower()}
        print(f"  Filtered to: {len(merged)} vendors")

    if not merged:
        print("\n[OK] No changes detected")
        return results

    # Step 2.5: Analyze quote attachments
    quote_analysis = {}
    has_quote_changes = any(
        v.get("quote_count", 0) > 0 for v in merged.values()
    )

    if has_quote_changes and AttachmentDownloader is not None:
        print("\n[Step 2.5] Analyzing quote attachments...")
        try:
            analyzer = QuoteAnalyzer(gmail_client=gmail_client)
            quote_analysis = analyzer.analyze_quote_emails(
                gmail_formatted, gmail_changes
            )

            for vendor, analysis in quote_analysis.items():
                files_count = analysis.get("files_analyzed", 0)
                quote_summary = analysis.get("quote_summary", "N/A")
                error_count = len(analysis.get("errors", []))
                print(f"  {vendor}:")
                print(f"    Files: {files_count}")
                print(f"    Quotes: {quote_summary}")
                if error_count:
                    print(f"    Errors: {error_count}")
                    for err in analysis.get("errors", []):
                        print(f"      - {err}")

            results["quote_analysis"] = {
                vendor: {
                    "files_analyzed": a.get("files_analyzed", 0),
                    "quote_summary": a.get("quote_summary", "N/A"),
                    "options_count": len(a.get("options", [])),
                }
                for vendor, a in quote_analysis.items()
            }
        except Exception as e:
            print(f"  [ERR] Quote analysis: {e}")
            results["errors"].append(f"Quote analysis: {e}")
    elif has_quote_changes:
        print("\n[Step 2.5] Quote attachments detected but AttachmentDownloader not available")
        print("  Install dependencies: pip install pdfplumber openpyxl")

    # Step 3: Generate updates
    print("\n[Step 3] Generating updates...")

    updates = []
    negative_changes_skipped = []

    for vendor_name, changes in merged.items():
        update = {"vendor": vendor_name, "fields": {}}

        # Latest next_action
        if changes["next_actions"]:
            update["fields"]["next_action"] = changes["next_actions"][-1]

        # Latest last_contact
        if changes["last_contact"]:
            update["fields"]["last_contact"] = changes["last_contact"]

        # 견적 정보 통합: 첨부파일 분석 결과가 있으면 quote 필드 업데이트
        if vendor_name in quote_analysis:
            va = quote_analysis[vendor_name]
            new_quote_summary = va.get("quote_summary", "")
            if new_quote_summary and new_quote_summary not in ("N/A", "금액 미감지"):
                # 현재 Slack Lists의 quote와 비교
                current = next(
                    (i for i in current_items if i["vendor_name"] == vendor_name),
                    None,
                )
                current_quote = current.get("quote", "") if current else ""

                # 첨부파일 분석 결과가 더 상세하면 업데이트 제안
                if len(new_quote_summary) > len(current_quote) or not current_quote.strip():
                    update["fields"]["quote"] = new_quote_summary
                    update["fields"]["quote_source"] = "attachment_analysis"

        # 견적 횟수 메타데이터 추가
        if changes.get("quote_count", 0) > 0:
            update["fields"]["quote_count"] = changes["quote_count"]

        # Status change (if any) - only apply POSITIVE changes automatically
        if changes["status_changes"]:
            latest = changes["status_changes"][-1]
            direction = latest.get("direction", "positive")

            if direction == "positive":
                # 상태 다운그레이드 방지: 현재 상태보다 낮은 상태로의 전환은 무시
                suggested_status = latest.get("to")
                current = next(
                    (i for i in current_items if i["vendor_name"] == vendor_name),
                    None,
                )

                if current:
                    current_status = current.get("status", "")
                    current_priority = STATUS_PRIORITY.get(current_status, 0)
                    suggested_priority = STATUS_PRIORITY.get(suggested_status, 0)

                    if suggested_priority <= current_priority and current_priority > 0:
                        print(f"  [SKIP] {vendor_name}: 상태 다운그레이드 방지 ({current_status} >= {suggested_status})")
                    else:
                        update["fields"]["status_suggestion"] = latest
                else:
                    update["fields"]["status_suggestion"] = latest
            else:
                # Negative status changes require explicit user approval
                negative_changes_skipped.append({
                    "vendor": vendor_name,
                    "change": latest,
                })
                print(f"  [SKIP] {vendor_name}: Negative status change skipped (requires manual approval)")
                print(f"         Trigger: '{latest.get('trigger')}' -> {latest.get('to')}")

        if update["fields"]:
            updates.append(update)
            # Safe print (handle encoding errors)
            try:
                vendor_display = vendor_name.encode('ascii', 'replace').decode()
            except Exception:
                vendor_display = vendor_name
            print(f"  * {vendor_display}:")
            for key, val in update["fields"].items():
                try:
                    val_display = str(val).encode('ascii', 'replace').decode()
                except Exception:
                    val_display = str(val)
                print(f"    - {key}: {val_display}")

    results["changes_detected"] = updates

    if dry_run:
        print("\n[DRY RUN] Changes NOT applied")
        return results

    # Step 4: Apply updates
    if not auto_approve:
        print("\n[SKIP] auto_approve=False - Changes NOT applied (user confirmation needed)")
        return results

    print("\n[Step 4] Applying updates...")

    for update in updates:
        vendor = update["vendor"]
        fields = update["fields"]

        try:
            # Parse last_contact date if present
            last_contact_dt = None
            if fields.get("last_contact"):
                lc = fields["last_contact"]
                if isinstance(lc, datetime):
                    last_contact_dt = lc
                elif isinstance(lc, str):
                    try:
                        from dateutil.parser import parse
                        last_contact_dt = parse(lc)
                    except Exception:
                        pass

            success = manager.update_item(
                vendor_name=vendor,
                quote=fields.get("quote"),
                next_action=fields.get("next_action"),
                last_contact=last_contact_dt,
            )
            if success:
                results["changes_applied"].append(update)
                print(f"  [OK] {vendor} updated")
            else:
                print(f"  [FAIL] {vendor} update failed")
        except Exception as e:
            print(f"  [ERR] {vendor}: {e}")
            results["errors"].append(f"{vendor}: {e}")

    # Step 5: Update summary
    print("\n[Step 5] Updating summary message...")
    try:
        updated_items = manager.get_list_items()
        summary = manager.generate_summary_message(updated_items)
        ts = manager.post_summary(summary)
        if ts:
            print(f"  [OK] Pinned message updated (ts: {ts})")
        else:
            print("  [FAIL] Pinned message update failed")
    except Exception as e:
        print(f"  [ERR] Summary update: {e}")
        results["errors"].append(f"Summary: {e}")

    # Save incremental state
    if incremental and not dry_run:
        sync_state.save()
        print("\n[Step 6] State saved to .slacklist_state.json")

    print(f"\n{'='*60}")
    print("[DONE] Complete")
    print(f"  Analyzed: Gmail {results['analyzed']['gmail']}, Slack {results['analyzed']['slack']}")
    print(f"  Detected: {len(results['changes_detected'])} changes")
    print(f"  Applied: {len(results['changes_applied'])} changes")
    if results["errors"]:
        print(f"  Errors: {len(results['errors'])}")
    print(f"{'='*60}\n")

    return results


def main():
    """CLI entry point."""
    import argparse

    parser = argparse.ArgumentParser(description="Sync Slack Lists")
    parser.add_argument(
        "--dashboard",
        type=Path,
        default=Path("C:/claude/wsoptv_ott/docs/management/VENDOR-DASHBOARD.md"),
        help="Path to VENDOR-DASHBOARD.md",
    )
    parser.add_argument(
        "--post-summary",
        action="store_true",
        help="Post summary to Slack channel",
    )
    parser.add_argument(
        "--intelligent-update",
        action="store_true",
        help="Intelligently update from Gmail and Slack analysis",
    )
    parser.add_argument(
        "--days",
        type=int,
        default=7,
        help="Number of days to analyze",
    )
    parser.add_argument(
        "--yes",
        action="store_true",
        help="Auto-approve changes without confirmation",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Show what would be done without making changes",
    )
    parser.add_argument(
        "--vendor",
        type=str,
        default=None,
        help="Filter to specific vendor (e.g., 'megazone', 'brightcove')",
    )
    parser.add_argument(
        "--full",
        action="store_true",
        help="Full scan mode (disable incremental, re-process all)",
    )
    parser.add_argument(
        "--status",
        action="store_true",
        help="Show sync state status",
    )

    args = parser.parse_args()

    if args.status:
        state = IncrementalSyncState()
        stats = state.get_stats()
        print("\n[SYNC STATE]")
        print(f"  Last sync: {stats['last_sync'] or 'Never'}")
        print(f"  Gmail processed: {stats['gmail_processed']}")
        print(f"  Slack processed: {stats['slack_processed']}")
        print(f"  Vendors tracked: {stats['vendors_tracked']}")
        print(f"  Pending changes: {stats['pending_changes']}")
    elif args.intelligent_update:
        intelligent_update_slacklist(
            days=args.days,
            dry_run=args.dry_run,
            auto_approve=args.yes,
            vendor_filter=args.vendor,
            incremental=not args.full,
        )
    elif args.post_summary:
        post_daily_summary(dry_run=args.dry_run)
    else:
        result = sync_lists_to_dashboard(
            dashboard_path=args.dashboard,
            dry_run=args.dry_run,
        )
        print(f"\n{result}")


if __name__ == "__main__":
    main()
